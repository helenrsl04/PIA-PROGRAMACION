contador = {   "Clientes": 0,                            
               "Salas": 0,                               
               "Programacion": 30
            }

cliente =  {   "CliCod": 0,
               "CliNam": 'Por Definir'              
            }


sala =     {   "sala 0": 'Sala Cero'       
           }

salaCapacidad =     {   "sala 0": '0'       
           }


agenda =   {   "0": '0'       
           }

#manejo de SQL Lite 3 
#20221029
import sys
import sqlite3
from sqlite3 import Error
#manejo de SQL Lite 3



import openpyxl                                                                 
import csv

def p_Creacion_Libro_Excel(fecha):
  libro = openpyxl.Workbook()
  hoja = libro["Sheet"] 

  hoja.title = "Reporte tabular"
  hoja['A1'].value='Reporte de reservaciones de Fecha = ' + fecha
  hoja['A2'].value=('Sala')
  #hoja['A3'].value= clave[8:9] #NombreSala
  hoja['B2'].value='Cliente'
  #hoja['B3'].value=clave[10:11] #NombreCliente
  hoja['C2'].value='Folio'
  #hoja['C3'].value=clave
  hoja['D2'].value='Evento'
  #hoja['D3'].value= agenda[clave] #AgendaEventoNombre
  hoja['E2'].value='Turno'
  #hoja['E3'].value=clave[9:10]

  v_renglon = 3
  for clave in agenda:
    if (len(clave) >= 10):
      if (clave[0:8] == fecha):
        #print(clave,"      ",clave[9:10],"      ",agenda[clave])

        #hoja.title = "Reporte tabular"
        #hoja['A1'].value='Reporte de reservaciones el dia '
        #hoja['A2'].value=('Sala')
        cellref=hoja.cell(v_renglon, column=1)
        cellref.value= clave[8:9] #NombreSala

        #hoja['B2'].value='Cliente'
        cellref=hoja.cell(v_renglon, column=2)
        cellref.value=clave[10:11] #NombreCliente
        #hoja['C2'].value='Folio'

        cellref=hoja.cell(v_renglon, column=3)
        cellref.value=clave
        #hoja['D2'].value='Evento'

        cellref=hoja.cell(v_renglon, column=4)
        cellref.value= agenda[clave] #AgendaEventoNombre
        #hoja['E2'].value='Turno'

        cellref=hoja.cell(v_renglon, column=5)
        cellref.value=clave[9:10]

        v_renglon = v_renglon + 1
  libro.save('ReporteTabular.xlsx')
  

def f_Encontro_Reservaciones_Fecha(fecha_a_buscar):
  v_contador = 0
  for clave in agenda:
      if (len(clave) >= 10):
        if (clave[0:8] == str(fecha_a_buscar)):
          v_contador = v_contador + 1
  if (v_contador > 0):
    return True
  else:
    return False
  

#solo valida que el string que intriuzca el usuario sea una Fecha Valida
def f_Fecha_NoValida(fecha_string):
  from datetime import datetime,date,timedelta
  # initializing format
  format = "%d-%m-%y"
  v_convierte_a_fecha = datetime.now()
  try:
    v_convierte_a_fecha = datetime.strptime(fecha_string, format)
    return False
  except ValueError:
    if (len(fecha_string) > 0):
      print("Fecha invalida ",fecha_string )
    return True

def f_EncontroCita(cita,agenda):                                                
   #### Validacion que no se repita Fecha-Sala-Turno 15-Oct-22                  
                  v_validacion = False                                          
                                                                                
                  for clave in agenda:                                          
                      if (len(clave) >= 10):                                    
                          if clave[0:10] == cita:                               
                             v_validacion = True                                

                  return v_validacion                                           
   #### Validacion que no se repita Fecha-Sala-Turno 15-Oct-22


def f_Error_En_Cadena(cadena):                                                
  if (len(cadena) == 0):
    print("No se permiten valores nulos.")
    print("")#Espacio
    return True
  elif ( len(cadena) != len(cadena.strip())):
    print("No se permiten espacios vacios.")
    print("")#espacio
    return True
  elif (not cadena.isalpha()) :
    print("El nombre no puede llevar numeros.")
    print("")#espacio
    return True
  else:                                                                        
    return False                                                               


#manejo de SQL Lite 3 
#20221030
def p_CreacionTablas():
  try:
      with sqlite3.connect("EvidenciaSQLlite.db") as conn:
          mi_cursor = conn.cursor()
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS t_contador (ConCod TEXT    , ConUlt INTEGER );")
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS t_cliente  (CliCod INTEGER , CliNom TEXT );")
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS t_sala     (SalCod INTEGER , SalNom TEXT ,SalCap TEXT);")          
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS t_agenda   (AgeCod TEXT ,    AgeNom TEXT);")                    

          mi_cursor.execute("DELETE FROM t_contador;")
          mi_cursor.execute("DELETE FROM t_cliente ;")
          mi_cursor.execute("DELETE FROM t_sala    ;")          
          mi_cursor.execute("DELETE FROM t_agenda  ;")                    


          #mi_cursor.execute("CREATE TABLE IF NOT EXISTS proyecto (clave INTEGER PRIMARY KEY, nombre TEXT NOT NULL, responsable INTEGER NOT NULL, FOREIGN KEY(responsable) REFERENCES responsable(clave));")
          print("Tablas creadas exitosamente")
  except Error as e:
      print (e)
  except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
  finally:
      conn.close()      

def f_SiguienteValorContador(valor_clave):
  #Lee todos los datos de una tabla que cumplan una condición
  #valor_clave = int(input("Clave del proyecto a consultar: "))

  try:
      with sqlite3.connect("EvidenciaSQLlite.db") as conn:
          mi_cursor = conn.cursor()
          valores = {"ConCod":valor_clave}
          mi_cursor.execute("SELECT * FROM t_contador WHERE ConCod = :ConCod", valores)
          registro = mi_cursor.fetchall()

          if registro:
              for ConCod, ConUlt in registro:
                  #print(f"{clave}\t{nombre}")
                  #Incrementa el contador
                  return ConUlt
          else:
              #print(f"No se encontró un proyecto asociado con la clave {valor_clave}")
              #Inserta el registro con valor 1 y devuelve 1
              return 0

  except Error as e:
      print (e)
  except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")  


def f_ValidaValorEnTabla(valor_clave):
  try:
      with sqlite3.connect("EvidenciaSQLlite.db") as conn:
          mi_cursor = conn.cursor()
          valores = {"ConCod":valor_clave}
          mi_cursor.execute("SELECT * FROM t_contador WHERE ConCod = :ConCod", valores)
          registro = mi_cursor.fetchall()

          if registro:
              for ConCod, ConUlt in registro:
                  #print(f"{clave}\t{nombre}")
                  #Incrementa el contador
                  return ConUlt
          else:
              #print(f"No se encontró un proyecto asociado con la clave {valor_clave}")
              #Inserta el registro con valor 1 y devuelve 1
              return 0

  except Error as e:
      print (e)
  except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")  



def p_EjecutaComandoSQL(SentenciaSql,Parametros):
  try:
      with sqlite3.connect("EvidenciaSQLlite.db") as conn:
          mi_cursor = conn.cursor()
          mi_cursor.execute(SentenciaSql, Parametros)
          #registros = mi_cursor.fetchall()
          #if registros:
            #return len(registros)
          #else:
            #return 0
          #print("SQL Ejecutado _Exitosamente") #################################
          print("") #Espacio
  except Error as e:
      print("Mensaje Error")
      print (e)
  except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
  finally:
      conn.close()  
#manejo de SQL Lite 3 


#  ************* MENU PRINCIPAL ****************
p_CreacionTablas()

opcion = '0'                                                                    
print("") #Espacio
while not(opcion=='5'):                                                         
    print("      **** MENU PRINCIPAL ****  ")
    print(' [1]. Reservaciones ')
    print(' [2]. Reportes')
    print(' [3]. Registrar una sala')
    print(' [4]. Registrar un cliente')
    print(' [5]. Salir')
    print("")

    opcion=input('  --- ¿Cuál opcion?: ')
#  ************* MENU PRINCIPAL ****************

#  ************* [1]. RESERVACIONES ****************
#  ************* SUBMENU 1. RESERVACIONES ****************
    if (opcion=='1'):                                                           
        opcion = '0'                                                            

        while not(opcion=='9'):
            print('          *** SUBMENU DE RESERVACIONES***             ')
            print(' [1]. Registrar una nueva reservación ')                     
            print(' [2]. Modificar descripción de una reservación ') 
            print(' [3]. Consultar disponibilidad de salas para una fecha ')
            print(' [4]. Eliminar una reservación') #Evidencia 3 nuevo
            print(' [5]. Salir')                                                
            print("") #espacio
#  ************* SUBMENU 1. RESERVACIONES ****************
            opcion=input('  --- ¿Cuál opcion?: ')


#  ************* SUBMENU 1. [1]. Registrar una nueva reservación ****************              
            if (opcion=='1'):
                print('Registrar una nueva reservación')

                #print(sala)
                #for clave in sala:
                  #print(clave)
                  #print(sala[clave])

                print("Sala    Descripcion")
                for key, value in sala.items():
                  if (key != "sala 0"):
                    print(key, '     ', value)




                ########## Validacion SALA #########
                while True:                                                     
                  try:                                                          
                    iSala = int(input("¿Numero de sala?: "))                              
                  except ValueError:                                            
                    print("") #Espacio
                    print("Debes escribir un número.")
                    print("")#espacio
                    continue                                                    

                  if (iSala < 0):                                               
                    print("Debes escribir un número positivo.")
                    print("")#espacio          
                    continue                                                    
                  else:                                                         
                    #break
                    if (iSala in sala):                                         
                      validar = sala.get(iSala)                                 
                      print(validar)
                      break
                    else:
                      print("No existe la sala " , iSala)
                      print("")#espacio
                      continue
                ########## Validacion SALA #########


                print("Cliente    Nombre")
                for key, value in cliente.items():
                  if (key != "CliCod" and key != "CliNam"):
                    print(key, '     ', value)




                ########## Validacion CLIENTE #########
                while True:
                  try:
                    iCliente = int(input("¿Numero de cliente?: "))
                  except ValueError:
                    print("Debes escribir un número.")
                    print("")#espacio
                    continue

                  if (iCliente < 0):
                    print("Debes escribir un número positivo.")
                    print("") #espacio
                    continue
                  else:
                    #break
                    if (iCliente in cliente):
                      validar = cliente.get(iCliente)
                      print(validar)
                      print("") #Espacio
                      break
                    else:
                      print("No existe el cliente " , iCliente)
                      continue
                ########## Validacion CLIENTE #########


                ########## Validacion TURNO #########
                while True:
                    iTurno = input("TURNO [M]añana [T]arde  [N]oche ?: ")

                    iTurno = iTurno.upper()

                    if (iTurno != "M" and iTurno != "T" and iTurno != "N"):
                      print("Turno no valido, debe ser [M] [T] [N]")
                      continue
                    else:
                      break
                ########## Validacion TURNO #########
                      

                ########## Validacion FECHA #########
                while True:
                    from datetime import datetime,date,timedelta

                    test_str = input("Fecha que desea hacer la reservación? (DD-MM-YY): ")
 
                    # initializing format
                    format = "%d-%m-%y"
                    EndDate = datetime.now()+timedelta(days=2)
                    #res = datetime.now()

                    # using try-except to check for truth value
                    try:
                       # res = bool(datetime.strptime(test_str, format))
                        res = datetime.strptime(test_str, format)
                        v_abuscar = test_str + str(iSala) + iTurno 
                        #break
                        if (res >= EndDate) and not f_EncontroCita(v_abuscar,agenda):
                          break
                        else:
                              if not (res >= EndDate):
                               print("La fecha debe ser con 2 dias de anticipación.")
                               continue
                              else:
                                print("Se encontro una reservación para ese dia, favor de introducir otra.")
                                continue
                    except ValueError:
                      #res = False
                      print("Fecha Invalida ",test_str )
                      continue
                ########## Validacion FECHA #########


                #Se genera consecutivo de agenda
                #8 Car fecha + 1 Car Sala +  1 Car Turno + 1 Car cliente
                AgendaFolio = test_str + str(iSala) + iTurno + str(iCliente) 
                print("Folio de agenda = ",AgendaFolio)

                AgendaEventoNombre = input("¿Nombre del evento?: ")
                agenda[AgendaFolio] = AgendaEventoNombre

                #manejo de SQL Lite 3 
                #20221031
                valores = (AgendaFolio, AgendaEventoNombre)
                sqlite_query = "INSERT INTO t_agenda VALUES(?,?)"
                p_EjecutaComandoSQL(sqlite_query ,valores)   
                #manejo de SQL Lite 3 
                print("")#Espacio
#  ************* SUBMENU 1. [1]. Registrar una nueva reservación **************** 
                
                
#  ************* SUBMENU 1. [2]. Modificar descripción de una reservación ****************                
                print("") #espacio
            elif (opcion=='2'):
                print('Modificar descripción de una reservación')
                print("") #espacio

                #Desplegado de Folios
                print("Folio         Evento")
                for key,value in agenda.items():
                  if (len(key) >= 10):
                    print(key,"      ",value)


                v_folio = input("Ingrese el folio de la reservación a modificar: ")
                v_folio = v_folio.strip()

                v_loencontre = False
                v_evento = ""
                v_eventomodificado = ""

                

                #if (v_Pregunta_Modificar == 'S'):
                for clave in agenda:
                   if (len(clave) >= 10):
                     if (clave == str(v_folio)):
                        #print(clave,"      ",clave[9:10],"      ",agenda[clave])
                        v_loencontre = True
                        v_evento = agenda[clave]

                if v_loencontre:
                    #v_eventomodificado = input("Para el Folio ",v_folio, "con _Descripcion " , v_evento , " Dame la nueva descripcion ?")
                    v_eventomodificado = input("Introduzca la nueva descripción: ")
                    agenda[clave] = v_eventomodificado

                    #manejo de SQL Lite 3 
                    #20221031
                    valores = (v_eventomodificado, v_folio)
                    sqlite_query = """Update t_agenda set AgeNom = ? where Agecod = ?"""
                    p_EjecutaComandoSQL(sqlite_query ,valores)          
                    #manejo de SQL Lite 3 
                    print("La modificación se realizo correctamente")
                else:
                  print("No existe el folio ", v_folio)
#  ************* SUBMENU 1. [2]. Modificiar descripción de una reservación **************** 


#  ************* SUBMENU 1. [3]. Consultar disponibilidad de salas para una fecha **************** 
            elif (opcion=='3'):
                print('Consultar disponibilidad de salas para una fecha')
                print("") #espacio

                v_fecha = ""
                while f_Fecha_NoValida(v_fecha):
                 v_fecha  = input("Introduzca la fecha a consultar disponibilidad: ")

                print("Sala" , " Turno")

                ##Recorer la Sala
                #for clave_sala in sala:
                for key,value in sala.items():
                  v_encontre_M = False
                  v_encontre_T = False
                  v_encontre_N = False
                  if (key != "sala 0"):  
                      for clave_agenda in agenda:
                        if (len(clave_agenda) >= 10):
                            #Para la Sala del ciclo se revisa la fecha , la Sala y el Turno M
                          if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(key) and (clave_agenda[9:10] == "M"  ):                 
                              v_encontre_M = True

                          if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(key) and (clave_agenda[9:10] == "T"  ):                 
                              v_encontre_T = True

                          #if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(clave_sala) and (clave_agenda[9:10] == "N"  ):                 
                          if (clave_agenda[0:8]) == v_fecha and (clave_agenda[8:9]) == str(key) and (clave_agenda[9:10] == "N"  ):                 
                              v_encontre_N = True

                      if not (v_encontre_M):
                        #print("Para fecha = ",v_fecha,"  y sala   = ",clave_sala , " Esta disponible Turno M")
                        print(value , "   M")
                          
                      if not (v_encontre_T):
                        print(value , "   T")

                      if not (v_encontre_N):
                        #print(clave_sala , "   N")
                        print(value , "   N")
#  ************* SUBMENU 1. [3]. Consultar disponibilidad de salas para una fecha **************** 


#  ************* SUBMENU 1. [4]. Eliminar una reservación **************** 
            elif (opcion=='4'):
                print('Borrado de una reservación')
                print("") #espacio

                v_loencontre = False
                v_evento = ""
                v_eventomodificado = ""
                v_folio = ""

                #Desplegado de Folios
                print("Folio         Evento")
                for key,value in agenda.items():
                  if (len(key) >= 10):
                    print(key,"      ",value)


                #Validacion Fecha A bORRAR
                format = "%d-%m-%y"
                EndDate = datetime.now()+timedelta(days=3)
                #while True:
                v_folio = input("Introduzca el folio de la agenda a borrar: ")
                v_folio = v_folio.strip()
                try:
                   v_solo_fecha =v_folio[0:8]
                   res = datetime.strptime(v_solo_fecha, format)
                   if (res >= EndDate) :
                     print("")
                     #break
                   else:
                     print("La fecha no es mayor a la fecha actual + 3 dias")
                     #continue
                except ValueError:
                   #res = False
                   print("Fecha Invalida ",v_solo_fecha )
                   print("")
                   #continue
                #Validacion Fecha A bORRAR

                v_Pregunta_Borrar = ""
                while (v_Pregunta_Borrar.upper() != "S" and v_Pregunta_Borrar.upper() != "N" ):
                  v_Pregunta_Borrar = input("¿Esta seguro que desea borrar el folio S/N? ")
                  v_Pregunta_Borrar = v_Pregunta_Borrar.upper()

                if (v_Pregunta_Borrar == 'S'):
                  for clave in agenda:
                    if (len(clave) >= 10):
                      if (clave == str(v_folio)):
                          #print(clave,"      ",clave[9:10],"      ",agenda[clave])
                          v_loencontre = True
                          v_evento = agenda[clave]

                  if v_loencontre:
                      print("Borrando el folio =",agenda[clave])
                      del agenda[v_folio]                
                      try:
                        sqliteConnection = sqlite3.connect('EvidenciaSQLlite.db')
                        cursor = sqliteConnection.cursor()
                        #print("Connected to SQLite")

                        sql_update_query = """DELETE from t_agenda where AgeCod = ?"""
                        cursor.execute(sql_update_query, (v_folio,))
                        sqliteConnection.commit()
                        #print("Record deleted successfully") #############
                        cursor.close()
                      except sqlite3.Error as error:
                        print("Failed to delete reocord from a sqlite table", error)
                      finally:
                        if sqliteConnection:
                            sqliteConnection.close()

                      print("Se borro el Folio: ",v_folio)
                      print("") #Espacio
                ########## Validacion FECHA #########
#  ************* SUBMENU 1. [4]. Eliminar una reservación **************** 


#  ************* SUBMENU 1. [5]. Saliendo del submenu de reservaciones **************** 
            elif (opcion=='5'):
                print(' ** Saliendo del submenu de reservaciones **')
                print("")#espacio
                opcion = '0'
                break
            else:
                print('No existe la opcion.')
        print("") #Espacio
#  ************* SUBMENU 1. [5]. Saliendo del submenu de reservaciones **************** 
#  ************* SUBMENU 1. RESERVACIONES ****************


#  ************* SUBMENU 2. REPORTES ****************
    if (opcion=='2'):
        opcion = '0' 

        while not(opcion=='9'):
            print('          *** SUBMENU DE REPORTES***             ')
            print(' [1]. Reporte en pantalla de reservaciones para una fecha')                 
            print(' [2]. Exportar reporte tabular en Excel') 
            print(' [3]. Salir') # en todas son 5 pero se veria mal?                                                
            print("") #espacio

            opcion=input('  --- ¿Cuál opcion?: ')
        
#  ************* SUBMENU 2. [1]. Reporte en pantalla de reservaciones para una fecha ****************         
            if (opcion=='1'):
                print('[1]. Reporte en pantalla de reservaciones para una fecha')
                print("") #espacio
    

                v_fecha = ""
                while f_Fecha_NoValida(v_fecha):
                  v_fecha  = input("¿Qué fecha desea ver sus reservaciones?:  ")

                if f_Encontro_Reservaciones_Fecha(v_fecha):
                  print('*****************************************************')
                  print(' **** Reporte de reservaciones el dia ',v_fecha,'****')
                  print('Folio            Turno       Nombre Evento')
                  print('*****************************************************')
                  for clave in agenda:
                      if (len(clave) >= 10):
                        if (clave[0:8] == str(v_fecha)):
                          print(clave,"      ",clave[9:10],"      ",agenda[clave])
                  print(' ****************** Fin del reporte *****************')
                else:
                  print("No existe reservaciones para la fecha: ",v_fecha)
                print("") #Espacio
#  ************* SUBMENU 2. [1]. Reporte en pantalla de reservaciones para una fecha ****************               


#  ************* SUBMENU 2. [2]. Exportar reporte tabular en Excel **************** 
            elif (opcion=='2'):
                print('[2].Exportar reporte tabular en Excel')
                print("") #espacio

                v_fecha = ""
                while f_Fecha_NoValida(v_fecha):
                  v_fecha  = input("¿Qué fecha desea exportar al Excel?:  ")

                if f_Encontro_Reservaciones_Fecha(v_fecha):
                  print("Creando Archivo de Excel para Fecha: ",v_fecha)
                  p_Creacion_Libro_Excel(v_fecha)
                  print('El libro de excel fue creado exitosamente para fecha:',v_fecha)
                  print("")#espacio  
                else:
                  print("No existe reservaciones para la fecha: ",v_fecha)
                  print("")#espacio

#  ************* SUBMENU 2. [2]. Exportar reporte tabular en Excel **************** 


#  ************* SUBMENU 2. [3]. Saliendo del submenu de reportes **************** 
            elif (opcion=='3'):
                print(' **  Saliendo del submenu de reportes  **')
                print("") #Espacio
                opcion = '0'
                break
            else:
                print('No existe la opcion..')
                print("") #Espacio
#  ************* SUBMENU 2. [3]. Saliendo del submenu de reportes ****************
#  ************* [1]. RESERVACIONES ****************


#  ************* [4]. REGISTAR UN CLIENTES ****************
    elif (opcion=='4'):
        print(' **** Cliente ****')
        #print(cliente) Andy 15-10-22 
        #print("") #Espacio ################################################

        #manejo de SQL Lite 3 
        #20221029
        siguiente = f_SiguienteValorContador('Clientes')
        #print("Debug siguiente llamado a l a funcion,") ##############
        #print(siguiente) ###########

        #manejo de SQL Lite 3 

        siguiente =  siguiente + 1
        #print("Debug siguiente sumado") ##########
        #print(siguiente) ################


        if (siguiente == 1):
          #Inserta el Nuevo contador
          #manejo de SQL Lite 3 
          #20221031
          valores = ('Clientes', siguiente)
          sqlite_query = "INSERT INTO t_contador VALUES(?,?)"
          p_EjecutaComandoSQL(sqlite_query ,valores)   
          #manejo de SQL Lite 3 
    

        else:
          ##Actualiza el contador
          # update value
          #siguiente =  siguiente + 1
          #contador['Clientes'] = siguiente

          #manejo de SQL Lite 3 
          #20221031
          valores = (siguiente,  'Clientes')
          sqlite_query = """Update t_contador set ConUlt = ? where ConCod = ?"""
          p_EjecutaComandoSQL(sqlite_query ,valores)          
          #manejo de SQL Lite 3 

    
        print('Nuevo cliente',siguiente)

        
        NombreCliente = ""
        v_validacion_Error = True
        while v_validacion_Error:
          NombreCliente=input('¿Nombre del cliente?: ')
          if not f_Error_En_Cadena(NombreCliente):
            v_validacion_Error = False
            #print("Nose puede dejar vacio, favor de ingresar un nombre ")
 

        #sala['SalCod'] = siguiente
        #sala['SalNam'] = NombreCliente

        #sigue actualizando la estructura
        cliente[siguiente] = NombreCliente

        #manejo de SQL Lite 3 
        #20221031
        valores = (siguiente,NombreCliente)
        sqlite_query = "INSERT INTO t_cliente VALUES(?,?)"
        p_EjecutaComandoSQL(sqlite_query ,valores)          
        #manejo de SQL Lite 3         

        #print(siguiente)
        print("El cliente fue registrado correctamente") #Andy 15-10-22 Quitare lo que muestralas cadenas de cliente para que solo se vea que se ingreso correctamente
        print("") #Andy 15-10-22 Para que no se vea todo junto
#  ************* [4]. REGISTAR UN CLIENTES ****************


#  ************* [3]. REGISTAR UNA SALA ****************
    elif (opcion=='3'):
        print(' **** Salas ****')
        #print("") #Espacio
        #print(sala) Andy 15-10-22 Para que no se vean las listas al usuario
        
        #manejo de SQL Lite 3 
        #20221029
        siguiente = f_SiguienteValorContador('Salas')
        #print("Debug siguiente llamado a l a funcion,") ########
        #print(siguiente) ############

        #manejo de SQL Lite 3 

        siguiente =  siguiente + 1
        #print("Debug siguiente sumado") ######
        #print(siguiente)#####


        if (siguiente == 1):
          #Inserta el Nuevo contador
          #manejo de SQL Lite 3 
          #20221031
          valores = ('Salas', siguiente)
          sqlite_query = "INSERT INTO t_contador VALUES(?,?)"
          p_EjecutaComandoSQL(sqlite_query ,valores)          
          #manejo de SQL Lite 3 


        else:
          ##Actualiza el contador
          # update value
          #siguiente =  siguiente + 1
          #contador['Clientes'] = siguiente

          #manejo de SQL Lite 3 
          #20221031
          valores = (siguiente,  'Salas')
          sqlite_query = """Update t_contador set ConUlt = ? where ConCod = ?"""
          p_EjecutaComandoSQL(sqlite_query ,valores)          
          #manejo de SQL Lite 3 
  


        print('Nueva Sala',siguiente)
        

        NombreSala = ""
        v_validacion_Error = True
        while v_validacion_Error:
           NombreSala=input('¿Nombre de la sala?: ')
           if not f_Error_En_Cadena(NombreSala):
             v_validacion_Error = False



        CapacidadSala=0


        #####Validacion de Capacidad
        while True:
          try:
             CapacidadSala=int(input(f'¿Capacidad de la sala?: '))
             if  (CapacidadSala <= 0):
               print(f"Se de debe escribir un número mayor a 0 ")
               print("")#espacio
               #CapacidadSala=int(input('¿Capacidad Sala?: '))  
             else:
               break
          except ValueError:
            print("Favor de ingresar un dato valido")
            print("")#espacio
            continue
        #####Validacion de Capacidad  


        #sala['SalCod'] = siguiente
        #sala['SalNam'] = NombreCliente

        #Sigue Actualizando la Sala
        sala[siguiente] = NombreSala

        #salaCapacidad[siguiente] = CapacidadSala

        #manejo de SQL Lite 3 
        #20221031
        valores = (siguiente,NombreSala,CapacidadSala)
        sqlite_query = "INSERT INTO t_sala VALUES(?,?,?)"
        p_EjecutaComandoSQL(sqlite_query ,valores)          
        #manejo de SQL Lite 3 

        #print(siguiente)
        print("La sala fue creada correctamente") #Andy 15-10-22 
        #print(salaCapacidad) Andy 15-10-22 Esto era lo que mostraba las llaves pero no se ocupan para el usuario
        
        # Andy 16-10-22 La capacidad de la sala debe ser mayor a 0
        print("")#Espacio
#  ************* [3]. REGISTAR UNA SALA ****************


#  ************* [5]. SALIENDO DEL MENU ****************
    elif (opcion=='5'):
        print(' ** Saliendo del menu  **')
    #else:
        #print('No existe la opcion...')